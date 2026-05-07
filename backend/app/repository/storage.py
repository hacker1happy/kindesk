import json
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[3]
DATA_DIR = PROJECT_ROOT / "data"
UPLOADS_DIR = DATA_DIR / "uploads"

CLIENTS_DB_PATH = DATA_DIR / "clients.json"
CASES_DB_PATH = DATA_DIR / "cases.json"
COMPANIES_DB_PATH = DATA_DIR / "companies_master.json"
RTAS_DB_PATH = DATA_DIR / "rta_master.json"
MASTER_WORKBOOK_PATH = DATA_DIR / "kindesk_companies.xlsx"


def read_json(path):
    path = Path(path)
    if not path.exists():
        return {}
    with open(path, "r") as f:
        return json.load(f)


def write_json(path, data):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w") as f:
        json.dump(data, f, indent=2)


def read_clients():
    return read_json(CLIENTS_DB_PATH)


def write_clients(data):
    write_json(CLIENTS_DB_PATH, data)


def read_cases():
    return read_json(CASES_DB_PATH)


def write_cases(data):
    write_json(CASES_DB_PATH, data)


def read_workbook_sheet(sheet_name: str, key_column: str):
    if not MASTER_WORKBOOK_PATH.exists():
        return {}

    workbook = load_workbook(MASTER_WORKBOOK_PATH, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return {}

        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(rows, [])]
        records = {}

        for row in rows:
            item = {
                headers[index]: value
                for index, value in enumerate(row)
                if index < len(headers) and headers[index]
            }
            record_key = item.get(key_column)
            if not record_key:
                continue

            records[str(record_key).strip()] = {
                key: str(value).strip() if value is not None else ""
                for key, value in item.items()
            }

        return records
    finally:
        workbook.close()


def read_companies():
    return read_workbook_sheet("companies_master", "company_id")


def read_rtas():
    return read_workbook_sheet("rta_master", "rta_id")


def read_data():
    return read_clients()


def write_data(data):
    write_clients(data)
